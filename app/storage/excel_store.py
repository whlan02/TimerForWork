from __future__ import annotations

from dataclasses import asdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook

from app.models.record import TimeRecord


class ExcelStore:
    def __init__(self, file_path: Optional[Path] = None) -> None:
        self.file_path: Path = Path(file_path) if file_path else Path.cwd() / "time_records.xlsx"
        self.sheet_name: str = "records"
        self.date_to_records: Dict[date, List[TimeRecord]] = {}
        self.date_to_total_minutes: Dict[date, int] = {}
        self._ensure_file()
        self._load_all()

    def _ensure_file(self) -> None:
        if not self.file_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["date", "start_time", "end_time", "duration_sec", "duration_min", "note"])
            wb.save(self.file_path)
            wb.close()
        else:
            wb = load_workbook(self.file_path)
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(self.sheet_name)
                ws.append(["date", "start_time", "end_time", "duration_sec", "duration_min", "note"])
                wb.save(self.file_path)
            else:
                # 兼容旧文件：若缺少 duration_sec 列则补上
                ws = wb[self.sheet_name]
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                if header and "duration_sec" not in header:
                    # 在 duration_min 前或末尾插入一列标题（为简单起见，追加到末尾）
                    ws.cell(row=1, column=len(header) + 1, value="duration_sec")
                wb.save(self.file_path)
            wb.close()

    def _load_all(self) -> None:
        self.date_to_records.clear()
        self.date_to_total_minutes.clear()
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]
        # 读取表头映射，支持旧/新列顺序
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(h) if h is not None else "" for h in header_cells]
        name_to_idx = {name: idx for idx, name in enumerate(header)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            # 兼容旧行：从列名安全取值
            d_str = row[name_to_idx.get("date", -1)] if name_to_idx.get("date", -1) >= 0 else None
            start_str = row[name_to_idx.get("start_time", -1)] if name_to_idx.get("start_time", -1) >= 0 else None
            end_str = row[name_to_idx.get("end_time", -1)] if name_to_idx.get("end_time", -1) >= 0 else None
            duration_min = row[name_to_idx.get("duration_min", -1)] if name_to_idx.get("duration_min", -1) >= 0 else None
            duration_sec = row[name_to_idx.get("duration_sec", -1)] if name_to_idx.get("duration_sec", -1) >= 0 else None
            note = row[name_to_idx.get("note", -1)] if name_to_idx.get("note", -1) >= 0 else ""
            if d_str is None or start_str is None or end_str is None:
                continue
            try:
                d: date = d_str if isinstance(d_str, date) else datetime.strptime(str(d_str), "%Y-%m-%d").date()
                st: time = (
                    start_str
                    if isinstance(start_str, time)
                    else datetime.strptime(str(start_str), "%H:%M:%S").time()
                )
                et: time = (
                    end_str
                    if isinstance(end_str, time)
                    else datetime.strptime(str(end_str), "%H:%M:%S").time()
                )
                mins: int = int(duration_min) if duration_min is not None else 0
                secs: int = int(duration_sec) if duration_sec is not None else int(mins * 60)
                rec = TimeRecord(date=d, start_time=st, end_time=et, duration_min=mins, duration_sec=secs, note=str(note or ""))
            except Exception:
                # skip malformed rows
                continue
            self.date_to_records.setdefault(rec.date, []).append(rec)
            self.date_to_total_minutes[rec.date] = self.date_to_total_minutes.get(rec.date, 0) + rec.duration_min
        wb.close()

    def get_records_for_date(self, d: date) -> List[TimeRecord]:
        return list(self.date_to_records.get(d, []))

    def get_total_minutes(self, d: date) -> int:
        return int(self.date_to_total_minutes.get(d, 0))

    def add_record(self, record: TimeRecord) -> None:
        # append to excel
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]
        # 确保包含 duration_sec 列；若无则追加表头列
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "duration_sec" not in header:
            ws.cell(row=1, column=len(header) + 1, value="duration_sec")
            header.append("duration_sec")
        # 以当前表头顺序写入一行，未知列留空
        row_data_map = {
            "date": record.date.strftime("%Y-%m-%d"),
            "start_time": record.start_time.strftime("%H:%M:%S"),
            "end_time": record.end_time.strftime("%H:%M:%S"),
            "duration_sec": int(record.duration_sec),
            "duration_min": int(record.duration_min),
            "note": record.note or "",
        }
        row_to_write = [row_data_map.get(str(col_name), "") for col_name in header]
        ws.append(row_to_write)
        wb.save(self.file_path)
        wb.close()
        # update in-memory indexes
        self.date_to_records.setdefault(record.date, []).append(record)
        self.date_to_total_minutes[record.date] = self.date_to_total_minutes.get(record.date, 0) + record.duration_min



